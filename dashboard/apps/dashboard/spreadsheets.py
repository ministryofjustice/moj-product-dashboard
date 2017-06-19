# -*- coding: utf-8 -*-
from datetime import date

from openpyxl.styles import Style, Font
from openpyxl.workbook import Workbook

from dashboard.libs.date_tools import parse_date
from .models import Person


CURRENCY_FORMAT = '£#,##0.00'


class Products:

    date_style = Style(number_format='DD/MM/YYYY')
    header_style = Style(font=Font(bold=True))
    currency_style = Style(number_format='£#,##0.00')

    def __init__(self, products, calculation_start_date=None):
        self.workbook = Workbook()
        self.fill_main_sheet(self.workbook.active, products, calculation_start_date)
        # create the monthly spend sheet when there is one product.
        # there is no use case for monthly spend sheet for multiple
        # products. if it's needed, column alignment needs to be
        # implemented.
        if len(products) == 1:
            self.fill_monthly_spend_sheet(
                self.workbook.create_sheet(), products[0], calculation_start_date)

    def fill_product_info(self, product, sheet, row, calculation_start_date, include_header=False):

        fields = [
            # (header, style, value)
            ('Id', None, product.id),
            ('Name', None, product.name),
            ('Description', None, product.description),
            ('Area name', None, product.area_name),
            ('Discovery date', self.date_style, product.discovery_date),
            ('Alpha_date', self.date_style, product.alpha_date),
            ('Beta date', self.date_style, product.beta_date),
            ('Live date', self.date_style, product.live_date),
            ('End date', self.date_style, product.end_date),
            ('Discovery fte', None, product.discovery_fte),
            ('Alpha fte', None, product.alpha_fte),
            ('Beta fte', None, product.beta_fte),
            ('Live fte', None, product.live_fte),
            ('Final budget', self.currency_style, product.final_budget),
            ('Cost of discovery', self.currency_style,
             product.cost_of_discovery(calculation_start_date=calculation_start_date)),
            ('Cost of alpha', self.currency_style,
             product.cost_of_alpha(calculation_start_date=calculation_start_date)),
            ('Cost of beta', self.currency_style,
             product.cost_of_beta(calculation_start_date=calculation_start_date))
        ]

        def _financial_year_col(year):
            header = 'Cost in FY {}-{}'.format(str(year)[2:], str(year + 1)[2:])
            return header, self.currency_style, product.cost_in_fy(
                year, calculation_start_date=calculation_start_date)

        fields += [
            _financial_year_col(date.today().year + offset)
            for offset in range(-2, 2)
        ]

        fields += [
            ('Cost of sustaining', self.currency_style,
             product.cost_of_sustaining(calculation_start_date=calculation_start_date)),
            ('Total recurring costs', self.currency_style, product.total_recurring_costs),
            ('Savings enabled', self.currency_style, product.savings_enabled),
            ('Visible', None, product.visible),
        ]

        for idx, (header, style, value) in enumerate(fields):
            column = idx + 1
            if include_header:
                header_cell = sheet.cell(row=row, column=column)
                header_cell.style = self.header_style
                header_cell.value = header
                body_cell = sheet.cell(row=row + 1, column=column)
            else:
                body_cell = sheet.cell(row=row, column=column)
            body_cell.value = value
            if style:
                body_cell.style = style

    def fill_main_sheet(self, sheet, products, calculation_start_date):
        sheet.title = 'Products info'
        # set the freeze_panes attribute to 'A2',
        # row 1 i.e. header, will always be viewable,
        # no matter where the user scrolls in the spreadsheet
        sheet.freeze_panes = 'A2'

        # fill first product together with headers
        self.fill_product_info(
            products[0], sheet, 1, calculation_start_date, include_header=True)
        row = 3
        for product in products[1:]:
            self.fill_product_info(product, sheet, row, calculation_start_date)
            row += 1

    def fill_monthly_spend_sheet(self, sheet, product, calculation_start_date):
        sheet.title = 'Monthly spend'
        time_frames = product.profile(
            freq='MS', calculation_start_date=calculation_start_date
        )['financial']['time_frames']
        row_names = ['budget', 'savings', 'total', 'remaining']
        for idx, name in enumerate(row_names):
            row_header = sheet.cell(row=idx + 2, column=1)
            row_header.style = self.header_style
            row_header.value = name

        for idx, time_frame in enumerate(sorted(time_frames.keys())):
            column = idx + 2
            col_header = sheet.cell(row=1, column=column)
            col_header.style = self.header_style
            col_header.value = parse_date(time_frame.split('~')[0]).strftime('%b %y')

            for idx, row_name in enumerate(row_names):
                cell = sheet.cell(row=idx + 2, column=column)
                cell.style = self.currency_style
                cell.value = round(time_frames[time_frame][row_name], 0)


class Export():
    def __init__(self, cleaned_data, title='Export'):
        self.cleaned_data = cleaned_data
        self.title = title

    def export(self):
        wb = Workbook()
        self.write(wb)
        return wb

    def write(self, wb):
        bold_font = Font(bold=True)
        bold_style = Style(font=bold_font)
        currency_style = Style(number_format=CURRENCY_FORMAT)

        fields = (
            ('Name', 'name', {}, None),
            ('Type', 'type', {}, None),
            ('Current', 'is_current', {}, None),
            ('Rate', 'base_rate_on', {'on': self.cleaned_data['date']},
             currency_style),
            ('Rate Type', 'rate_type', {}, None),
            ('Applied Daily Rate', 'rate_on', {'on': self.cleaned_data['date']},
             currency_style),
        )

        sheet = wb.active
        sheet.title = self.title
        for col, (heading, f, kwargs, style) in enumerate(fields):
            cell = sheet.cell(row=1, column=col + 1)
            cell.style = bold_style
            cell.value = heading
        sheet.freeze_panes = sheet['A2']

        people = Person.objects.filter()

        for row, product in enumerate(people):
            for col, (heading, f, kwargs, style) in enumerate(fields):
                val = getattr(product, f)
                if callable(val):
                    val = val(**kwargs)
                cell = sheet.cell(row=row + 2, column=col + 1)
                if style:
                    cell.style = style
                cell.value = val
