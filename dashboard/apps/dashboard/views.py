from datetime import datetime, date
import re
from collections import OrderedDict

from django.shortcuts import render, redirect
from django.http import Http404, HttpResponse
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.views.generic import View
from openpyxl.styles import Style, Font
from openpyxl.workbook import Workbook
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import viewsets, generics

from dashboard.libs.date_tools import parse_date
from dashboard.libs import swagger_tools
from .models import Product, Area, ProductGroup, Person
from .tasks import sync_float
from .serializers import PersonSerializer, PersonProductSerializer


def _product_meta(request, product):
    meta = {
        'can_edit': product.can_user_change(request.user),
        'admin_url': request.build_absolute_uri(product.admin_url)
    }
    return meta


def product_html(request, id):
    if not id:
        id = Product.objects.visible().first().id
        return redirect(reverse(product_html, kwargs={'id': id}))
    try:
        Product.objects.visible().get(id=id)
    except (ValueError, Product.DoesNotExist):
        raise Http404
    return render(request, 'common.html')


@api_view(['GET'])
def product_json(request, id):
    """
    detail view of a single product
    """
    request_data = request.GET
    try:
        product = Product.objects.visible().get(id=id)
    except (ValueError, Product.DoesNotExist):
        error = 'cannot find product with id={}'.format(id)
        return Response({'error': error}, status=404)

    start_date = request_data.get('startDate')
    if start_date:
        start_date = parse_date(start_date)
    end_date = request_data.get('endDate')
    if end_date:
        end_date = parse_date(end_date)
    # get the profile of the product for each month
    profile = product.profile(
        start_date=start_date,
        end_date=end_date,
        freq='MS')
    meta = _product_meta(request, product)
    return Response({**profile, 'meta': meta})


def product_group_html(request, id):
    if not id:
        id = ProductGroup.objects.first().id
        return redirect(reverse(product_group_html, kwargs={'id': id}))
    try:
        ProductGroup.objects.get(id=id)
    except (ValueError, ProductGroup.DoesNotExist):
        raise Http404
    return render(request, 'common.html')


@api_view(['GET'])
def product_group_json(request, id):
    """
    detail view of a single product group
    """
    # TODO handle errors
    try:
        product_group = ProductGroup.objects.get(id=id)
    except (ValueError, ProductGroup.DoesNotExist):
        error = 'cannot find product group with id={}'.format(id)
        return Response({'error': error}, status=404)

    # get the profile of the product group for each month
    profile = product_group.profile(freq='MS')
    meta = _product_meta(request, product_group)
    return Response({**profile, 'meta': meta})


class PersonViewSet(viewsets.ReadOnlyModelViewSet):
    """
    List view of persons
    """
    queryset = Person.objects.all()
    serializer_class = PersonSerializer


@swagger_tools.additional_schema
class PersonProductListView(generics.ListAPIView):
    """
    List view of products the person(id={person_id}) spends time on
    in the time window defined by start date and end date.
    """

    schema = OrderedDict([
        ('start_date', {
            'name': 'start_date',
            'required': False,
            'location': 'query',
            'type': 'string',
            'description': 'start date',
         }),
        ('end_date', {
            'name': 'end_date',
            'required': False,
            'location': 'query',
            'type': 'string',
            'description': 'end date',
         }),
    ])
    serializer_class = PersonProductSerializer

    def get_serializer_class(self):
        return PersonProductSerializer

    def get_queryset(self):
        person = Person.objects.get(id=self.kwargs.get('person_id'))
        return person.products

    def get_serializer_context(self):
        context = super().get_serializer_context()
        start_date = self.request.query_params.get('start_date')
        if start_date:
            start_date = parse_date(start_date)
        end_date = self.request.query_params.get('end_date')
        if end_date:
            end_date = parse_date(end_date)
        return {
            'start_date': start_date,
            'end_date': end_date,
            'person': Person.objects.get(id=self.kwargs.get('person_id')),
            **context
        }


def service_html(request, id):
    if not id:
        id = Area.objects.filter(visible=True).first().id
        return redirect(reverse(service_html, kwargs={'id': id}))
    try:
        Area.objects.filter(visible=True).get(id=id)
    except (ValueError, Area.DoesNotExist):
        raise Http404
    return render(request, 'common.html')


@api_view(['GET'])
def service_json(request, id):
    """
    detail view of a single service area
    """
    try:
        area = Area.objects.filter(visible=True).get(id=id)
    except (ValueError, Area.DoesNotExist):
        error = 'cannot find service area with id={}'.format(id)
        return Response({'error': error}, status=404)
    # get the profile of the service
    return Response(area.profile())


def portfolio_html(request):
    return render(request, 'common.html', {'body_classes': 'portfolio'})


@api_view(['GET'])
def services_json(request):
    """
    list view of all service areas
    """
    result = [area.profile() for area in Area.objects.filter(visible=True)]
    return Response(result)


@login_required
@api_view(['POST'])
def sync_from_float(request):
    """
    sync data from Float.com
    """
    sync_float.delay()
    return Response({
        'status': 'STARTED'
    })


def kwarg_vals(kwargs):
    vals = []
    for k, v in kwargs.items():
        if k == 'year':
            v = '%s-%s' % (str(v)[2:], str(v + 1)[2:])
        vals.append(v)
    return vals


class PortfolioExportView(View):
    http_method_names = ['get']

    def format(self, value):
        if isinstance(value, date):
            return value.strftime('%d/%m/%Y')
        return value

    def get(self, request, *args, **kwargs):
        show = kwargs.get('show', 'visible')
        now = datetime.now()
        fname = '%s_%s_%s.xlsx' % (
            'ProductData',
            show,
            now.strftime('%Y-%m-%d_%H:%M:%S'))

        date_style = Style(number_format='DD/MM/YYYY')
        bold_font = Font(bold=True)
        bold_style = Style(font=bold_font)
        currency_style = Style(number_format='£#,##0.00')
        year = date.today().year

        fields = (
            # (header, style, method kwargs)
            ('Id', None, {}),
            ('Name', None, {}),
            ('Description', None, {}),
            ('Area name', None, {}),
            ('Discovery date', date_style, {}),
            ('Alpha_date', date_style, {}),
            ('Beta date', date_style, {}),
            ('Live date', date_style, {}),
            ('End date', date_style, {}),
            ('Discovery fte', None, {}),
            ('Alpha fte', None, {}),
            ('Beta fte', None, {}),
            ('Live fte', None, {}),
            ('Final budget', currency_style, {}),
            ('Cost of discovery', currency_style, {}),
            ('Cost of alpha', currency_style, {}),
            ('Cost of beta', currency_style, {}),
            ('Cost in FY', currency_style, {'year': year - 2}),
            ('Cost in FY', currency_style, {'year': year - 1}),
            ('Cost in FY', currency_style, {'year': year}),
            ('Cost in FY', currency_style, {'year': year + 1}),
            ('Cost of sustaining', currency_style, {}),
            ('Total recurring costs', currency_style, {}),
            ('Savings enabled', currency_style, {}),
            ('Visible', None, {}),
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Products info'
        for col, (f, style, kwargs) in enumerate(fields):
            cell = sheet.cell(row=1, column=col + 1)
            cell.style = bold_style
            cell.value = '%s %s' % (f, ' '.join(str(k) for k in kwarg_vals(kwargs)))
        sheet.freeze_panes = sheet['A2']

        if show == 'visible':
            products = Product.objects.visible()
        elif show == 'all':
            products = Product.objects.all()
        else:
            products = Product.objects.filter(pk=show)

        for row, product in enumerate(products):
            for col, (f, style, kwargs) in enumerate(fields):
                val = getattr(product, re.sub('[^0-9a-zA-Z]+', '_', f).lower())
                if callable(val):
                    val = val(**kwargs)
                val = self.format(val)
                cell = sheet.cell(row=row + 2, column=col + 1)
                if style:
                    cell.style = style
                cell.value = val

        response = HttpResponse(
            content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'attachment; filename=%s' \
                                          % fname
        workbook.save(response)
        return response
